import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def get_floor_name(team_id):
    tid = team_id.upper()
    if tid.startswith("NEX0") or tid.startswith("QTX0"):
        return "Ground Floor"
    elif tid.startswith("NEX1") or tid.startswith("QTX1"):
        return "First Floor"
    elif tid.startswith("NEX2") or tid.startswith("QTX2"):
        return "Second Floor"
    elif tid.startswith("NEX3") or tid.startswith("QTX3"):
        return "Online / Virtual"
    return "Main Venue"

def generate_excel_report(db_path, output_excel_path):
    with open(db_path, "r") as f:
        db = json.load(f)

    teams = db.get("teams", [])
    
    cyber_teams = sorted([t for t in teams if t.get("track") == "Cyber Security"], key=lambda x: x["id"])
    med_teams = sorted([t for t in teams if t.get("track") == "Med-Tech"], key=lambda x: x["id"])

    wb = openpyxl.Workbook()
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    header_border = Border(
        left=Side(style='medium', color='475569'),
        right=Side(style='medium', color='475569'),
        top=Side(style='medium', color='475569'),
        bottom=Side(style='medium', color='475569')
    )

    # Helper function to style a team sheet
    def build_team_sheet(ws, sheet_title, team_list, header_color, zebra_color, is_cyber=True):
        ws.views.sheetView[0].showGridLines = True
        
        # Title Block
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = f"QUANTEXA 2026 — {sheet_title.upper()} ({len(team_list)} TEAMS)"
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Subtitle Block
        ws.merge_cells('A2:I2')
        sub_cell = ws['A2']
        now_str = datetime.now().strftime("%B %d, %Y • %I:%M %p")
        sub_cell.value = f"Official Team Details • Generated on {now_str} • Total Registered Teams in Track: {len(team_list)}"
        sub_cell.font = Font(name="Calibri", size=10, italic=True, color="475569")
        sub_cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        # Empty row
        ws.row_dimensions[3].height = 10

        # Table Headers
        headers = [
            "S.No", "Team ID", "Team Name", "Team Leader Name", 
            "Contact Phone", "Team Size", "Assigned Venue / Floor", 
            "Track", "Submission Status"
        ]
        
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header_title
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = header_border

        ws.row_dimensions[4].height = 26

        # Populate Rows
        for row_idx, t in enumerate(team_list, start=5):
            venue = get_floor_name(t["id"])
            status = t.get("status", "In Progress")
            
            row_values = [
                row_idx - 4,
                t["id"],
                t["name"],
                t.get("leaderName", "N/A"),
                t.get("leaderPhone", "N/A"),
                t.get("membersCount", 4),
                venue,
                t.get("track", ""),
                status
            ]

            fill_color = zebra_color if (row_idx % 2 == 0) else "FFFFFF"

            for col_num, val in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.value = val
                cell.border = thin_border
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                # Alignments and specific formatting
                if col_num in [1, 2, 5, 6, 9]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                if col_num == 2:
                    cell.font = Font(name="Calibri", size=10, bold=True, color="6B21A8" if is_cyber else "047857")
                elif col_num == 3:
                    cell.font = Font(name="Calibri", size=10, bold=True, color="0F172A")
                else:
                    cell.font = Font(name="Calibri", size=10, color="1E293B")

            ws.row_dimensions[row_idx].height = 20

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in [1, 2]: continue
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 1. Cyber Security Sheet
    ws_cyber = wb.active
    ws_cyber.title = "Cyber Security Track"
    build_team_sheet(
        ws_cyber, 
        "Cyber Security Track Team Details", 
        cyber_teams, 
        header_color="581C87", 
        zebra_color="F3E8FF", 
        is_cyber=True
    )

    # 2. Med-Tech Sheet
    ws_med = wb.create_sheet(title="Med-Tech Track")
    build_team_sheet(
        ws_med, 
        "Med-Tech Track Team Details", 
        med_teams, 
        header_color="065F46", 
        zebra_color="ECFDF5", 
        is_cyber=False
    )

    # 3. Overview & Summary Sheet
    ws_summary = wb.create_sheet(title="Track Summary Overview")
    ws_summary.views.sheetView[0].showGridLines = True
    
    ws_summary.merge_cells('A1:F1')
    t_cell = ws_summary['A1']
    t_cell.value = "QUANTEXA 2026 — MASTER TRACK & VENUE SUMMARY"
    t_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t_cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 30

    sum_headers = ["Venue / Location", "ID Range", "Cyber Security", "Med-Tech", "Total Teams", "Track Split %"]
    for col_idx, h in enumerate(sum_headers, 1):
        c = ws_summary.cell(row=3, column=col_idx)
        c.value = h
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = header_border
    ws_summary.row_dimensions[3].height = 24

    floors = {}
    for t in teams:
        fl = get_floor_name(t["id"])
        if fl not in floors:
            floors[fl] = {"total": 0, "cyber": 0, "med": 0}
        floors[fl]["total"] += 1
        if t.get("track") == "Cyber Security":
            floors[fl]["cyber"] += 1
        else:
            floors[fl]["med"] += 1

    prefix_map = {
        "Ground Floor": "QTX0001 / NEX0001 - QTX0046 / NEX0046",
        "First Floor": "QTX1001 / NEX1001 - QTX1047 / NEX1047",
        "Second Floor": "QTX2001 / NEX2001 - QTX2058 / NEX2058",
        "Online / Virtual": "QTX3001 / NEX3001 - QTX3023 / NEX3023",
    }

    r_idx = 4
    for fl_name, counts in floors.items():
        row_vals = [
            fl_name,
            prefix_map.get(fl_name, "N/A"),
            counts["cyber"],
            counts["med"],
            counts["total"],
            f"{(counts['total']/len(teams))*100:.1f}%"
        ]
        for col_idx, val in enumerate(row_vals, 1):
            c = ws_summary.cell(row=r_idx, column=col_idx)
            c.value = val
            c.border = thin_border
            c.font = Font(name="Calibri", size=10, color="1E293B")
            if col_idx in [3, 4, 5, 6]:
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        ws_summary.row_dimensions[r_idx].height = 20
        r_idx += 1

    # Total Row
    total_vals = ["GRAND TOTAL", "ALL VENUES", len(cyber_teams), len(med_teams), len(teams), "100.0%"]
    for col_idx, val in enumerate(total_vals, 1):
        c = ws_summary.cell(row=r_idx, column=col_idx)
        c.value = val
        c.font = Font(name="Calibri", size=10, bold=True, color="0F172A")
        c.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        c.border = thin_border
        if col_idx in [3, 4, 5, 6]:
            c.alignment = Alignment(horizontal="center", vertical="center")
        else:
            c.alignment = Alignment(horizontal="left", vertical="center")
    ws_summary.row_dimensions[r_idx].height = 22

    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 5, 14)

    wb.save(output_excel_path)
    print(f"Combined Excel report saved to: {output_excel_path}")

    # Build Standalone Med-Tech Excel File
    wb_med_only = openpyxl.Workbook()
    ws_m = wb_med_only.active
    ws_m.title = "Med-Tech Track"
    build_team_sheet(
        ws_m, 
        "Med-Tech Track Team Details", 
        med_teams, 
        header_color="065F46", 
        zebra_color="ECFDF5", 
        is_cyber=False
    )
    med_path = output_excel_path.replace("Quantexa_Admin_Track_Report.xlsx", "Quantexa_MedTech_Teams.xlsx")
    wb_med_only.save(med_path)
    print(f"Standalone Med-Tech Excel report saved to: {med_path}")

    # Build Standalone Cyber Security Excel File
    wb_cyber_only = openpyxl.Workbook()
    ws_c = wb_cyber_only.active
    ws_c.title = "Cyber Security Track"
    build_team_sheet(
        ws_c, 
        "Cyber Security Track Team Details", 
        cyber_teams, 
        header_color="581C87", 
        zebra_color="F3E8FF", 
        is_cyber=True
    )
    cyber_path = output_excel_path.replace("Quantexa_Admin_Track_Report.xlsx", "Quantexa_CyberSecurity_Teams.xlsx")
    wb_cyber_only.save(cyber_path)
    print(f"Standalone Cyber Security Excel report saved to: {cyber_path}")

if __name__ == "__main__":
    db_file = "./data/final_db.json"
    out_excel = "./public/uploads/Quantexa_Admin_Track_Report.xlsx"
    os.makedirs(os.path.dirname(out_excel), exist_ok=True)
    generate_excel_report(db_file, out_excel)

    # Copy to root directory as well
    generate_excel_report(db_file, "./Quantexa_Admin_Track_Report.xlsx")
